"""
Вспомогательные функции для генерации Excel-отчетов.
"""

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import send_file


# ── Стили ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True, size=10)
SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ── Создание книги ─────────────────────────────────────────────────

def create_workbook(title, subtitle, last_col='B', wrap=False):
    """Создать книгу с заголовком и подзаголовком.

    Возвращает (wb, ws, row, align_center, align_left).
    row — номер следующей свободной строки (после подзаголовка + пустая строка).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    align_c = CENTER_WRAP if wrap else CENTER
    align_l = LEFT_WRAP if wrap else LEFT

    ws[f'A1'] = title
    ws[f'A1'].font = TITLE_FONT
    ws.merge_cells(f'A1:{last_col}1')
    ws[f'A1'].alignment = align_c

    ws[f'A2'] = subtitle
    ws[f'A2'].font = SUBTITLE_FONT
    ws.merge_cells(f'A2:{last_col}2')

    return wb, ws, 4, align_c, align_l


# ── Секции и заголовки таблиц ──────────────────────────────────────

def write_section_header(ws, row, title, last_col='B'):
    """Записать заголовок секции (ПРОДУКТЫ / НОСИТЕЛИ / КОРОБКИ)."""
    ws[f'A{row}'] = title
    ws[f'A{row}'].font = SECTION_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    for col_letter in _col_range('A', last_col):
        ws[f'{col_letter}{row}'].fill = SECTION_FILL
    ws.merge_cells(f'A{row}:{last_col}{row}')
    return row + 1


def write_table_headers(ws, row, headers, align=CENTER):
    """Записать строку заголовков таблицы."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = align
        cell.border = BORDER
    return row + 1


# ── Строки данных ──────────────────────────────────────────────────

def write_data_row(ws, row, values, align_left=LEFT, align_center=CENTER,
                   left_columns=None):
    """Записать строку данных.

    left_columns — множество индексов (0-based), для которых выравнивание по левому краю.
    По умолчанию первый столбец выравнивается по левому, остальные — по центру.
    """
    if left_columns is None:
        left_columns = {0}
    for col_num, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = value
        cell.alignment = align_left if (col_num - 1) in left_columns else align_center
        cell.border = BORDER
    return row + 1


# ── Ширины колонок ─────────────────────────────────────────────────

def set_column_widths(ws, widths):
    """Установить ширины колонок. widths — dict {колонка: ширина}."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# ── Сохранение и отправка ──────────────────────────────────────────

def send_excel(wb, download_name):
    """Сохранить книгу в BytesIO и вернуть как файл для скачивания."""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name,
    )


# ── Утилиты ────────────────────────────────────────────────────────

def _col_range(start, end):
    """Вернуть список букв колонок от start до end включительно ('A'..'D')."""
    return [chr(c) for c in range(ord(start), ord(end) + 1)]


def now_str(fmt='%d.%m.%Y %H:%M:%S'):
    """Текущее время как строка."""
    return datetime.now().strftime(fmt)


def today_str(fmt='%Y%m%d'):
    """Текущая дата как строка."""
    return datetime.now().strftime(fmt)
