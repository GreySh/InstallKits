"""
Flask веб-приложение для Учета Инсталляционных Комплектов.
"""

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from datetime import datetime
import sys
import os
from io import BytesIO

# Добавляем путь к проекту
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from data import (
    add_product, get_all_products, get_product_by_id, get_product_components, update_product, delete_product,
    add_product_with_components, update_product_with_components,
    add_disc, get_all_discs, get_disc_by_id, get_disc_by_name, update_disc, delete_disc,
    add_box, get_all_boxes, get_box_by_id, get_box_by_name, update_box, delete_box,
    add_kit_component, update_kit_component, delete_kit_component, get_kit_components, get_kit_component,
    get_stock_disc_quantity, get_stock_box_quantity,
    adjust_stock_disc, adjust_stock_box,
    add_stock_disc, add_stock_box,
    get_all_stock_discs, get_all_stock_boxes,
    add_operation, get_operations, get_operations_by_type, delete_operation,
    get_product_available_quantity, dispatch_product, dispatch_products_batch
)
from settings_manager import get_db_path, set_db_path

# Импорт для Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Нужен для flash сообщений

# === Template helpers ===
@app.template_filter('default')
def default_filter(value, default=''):
    """Фильтр Jinja2 для установки значений по умолчанию."""
    return value if value is not None else default

# === Вспомогательные функции для шаблонов ===
def get_product_min_stock(product_id):
    """Получить минимальный остаток продукта."""
    return get_product_available_quantity(product_id)


def get_disc_count(product_id):
    """Получить количество носителей в продукте."""
    from data import get_product_components
    components = get_product_components(product_id)
    return len([c for c in components if c.get('disc_id')])


def get_box_count(product_id):
    """Получить количество коробок в продукте."""
    from data import get_product_components
    components = get_product_components(product_id)
    return len([c for c in components if c.get('box_id')])


# === Регистрация вспомогательных функций в Jinja2 ===
app.jinja_env.globals['get_product_min_stock'] = get_product_min_stock
app.jinja_env.globals['get_disc_count'] = get_disc_count
app.jinja_env.globals['get_box_count'] = get_box_count

# === Главная страница ===
@app.route('/')
def index():
    """Главная страница - Списание комплектов."""
    products = get_all_products()
    stock_discs = get_all_stock_discs()
    stock_boxes = get_all_stock_boxes()
    
    return render_template('view_stock.html', products=products, 
                          stock_discs=stock_discs, stock_boxes=stock_boxes)


# === Вкладка: Продукты ===
@app.route('/products')
def products():
    """Страница управления продуктами."""
    products = get_all_products()
    
    # Подсчитать количество компонентов для каждого продукта
    for product in products:
        components = get_product_components(product.doc_id)
        product.components_count = len(components) if components else 0
    
    return render_template('products.html', products=products)


@app.route('/products/add', methods=['GET', 'POST'])
def add_product_route():
    """Добавление нового продукта."""
    discs = get_all_discs()
    boxes = get_all_boxes()
    
    if request.method == 'POST':
        name = request.form.get('name')
        code = request.form.get('code')
        description = request.form.get('description')
        
        if name:
            # Подготовить список компонентов
            components = []
            for key, value in request.form.items():
                if key.startswith('component_disc_'):
                    disc_id = int(key.split('_')[-1])
                    quantity = int(request.form.get(f'quantity_disc_{disc_id}', 0))
                    if quantity > 0:
                        components.append({'disc_id': disc_id, 'disc_quantity': quantity, 'box_id': None, 'box_quantity': 0})
                elif key.startswith('component_box_'):
                    box_id = int(key.split('_')[-1])
                    quantity = int(request.form.get(f'quantity_box_{box_id}', 0))
                    if quantity > 0:
                        components.append({'disc_id': None, 'disc_quantity': 0, 'box_id': box_id, 'box_quantity': quantity})
            
            product_id = add_product_with_components(name, code, description, components)
            
            flash(f'Продукт "{name}" добавлен успешно!', 'success')
            return redirect(url_for('products'))
    
    return render_template('products_form.html', discs=discs, boxes=boxes, edit_mode=False)


@app.route('/products/edit/<int:product_id>', methods=['GET', 'POST'])
def edit_product_route(product_id):
    """Редактирование продукта."""
    product = get_product_by_id(product_id)
    discs = get_all_discs()
    boxes = get_all_boxes()
    current_components = get_product_components(product_id)
    
    # Подготовить данные для шаблона
    # kits use disc_id/box_id structure, not component_id
    current_components_discs = []
    current_components_boxes = []
    current_components_qty_disc = {}
    current_components_qty_box = {}
    
    if current_components:
        for comp in current_components:
            if comp.get('disc_id'):
                current_components_discs.append(comp['disc_id'])
                current_components_qty_disc[comp['disc_id']] = comp.get('disc_quantity', 0)
            if comp.get('box_id'):
                current_components_boxes.append(comp['box_id'])
                current_components_qty_box[comp['box_id']] = comp.get('box_quantity', 0)
    
    if request.method == 'POST':
        name = request.form.get('name')
        code = request.form.get('code')
        description = request.form.get('description')
        
        if name and product:
            # Подготовить новый список компонентов
            new_components = []
            
            # Сначала удаляем все существующие компоненты
            from data import delete_kit_component
            for comp in current_components:
                delete_kit_component(comp.doc_id)
            
            # Добавляем новые компоненты
            for key, value in request.form.items():
                if key.startswith('component_disc_'):
                    disc_id = int(key.split('_')[-1])
                    quantity = int(request.form.get(f'quantity_disc_{disc_id}', 0))
                    if quantity > 0:
                        new_components.append({'disc_id': disc_id, 'disc_quantity': quantity, 'box_id': None, 'box_quantity': 0})
                elif key.startswith('component_box_'):
                    box_id = int(key.split('_')[-1])
                    quantity = int(request.form.get(f'quantity_box_{box_id}', 0))
                    if quantity > 0:
                        new_components.append({'disc_id': None, 'disc_quantity': 0, 'box_id': box_id, 'box_quantity': quantity})
            
            update_product_with_components(product_id, name, code, description, new_components)
            
            flash(f'Продукт "{name}" обновлен успешно!', 'success')
            
            # Проверить откуда пришел запрос
            redirect_to = request.form.get('redirect_to')
            if redirect_to == 'kits':
                return redirect(url_for('kits'))
            else:
                return redirect(url_for('products'))
    
    return render_template('products_form.html', product=product, discs=discs, boxes=boxes, 
                          current_components_discs=current_components_discs,
                          current_components_boxes=current_components_boxes,
                          current_components_qty_disc=current_components_qty_disc,
                          current_components_qty_box=current_components_qty_box, edit_mode=True)


@app.route('/products/delete/<int:product_id>', methods=['POST'])
def delete_product_route(product_id):
    """Удаление продукта."""
    product = get_product_by_id(product_id)
    if product:
        delete_product(product_id)
        flash(f'Продукт "{product["name"]}" удален успешно!', 'warning')
    return redirect(url_for('products'))


# === Вкладка: Носители (Discs) ===
@app.route('/discs')
def discs():
    """Страница управления носителями."""
    discs = get_all_discs()
    return render_template('discs.html', discs=discs)


@app.route('/discs/add', methods=['POST'])
def add_disc_route():
    """Добавление нового носителя."""
    name = request.form.get('name')
    description = request.form.get('description')
    
    if name:
        add_disc(name, description)
        flash(f'Носитель "{name}" добавлен успешно!', 'success')
    
    return redirect(url_for('discs'))


@app.route('/discs/delete/<int:disc_id>', methods=['POST'])
def delete_disc_route(disc_id):
    """Удаление носителя."""
    disc = get_disc_by_id(disc_id)
    if disc:
        delete_disc(disc_id)
        flash(f'Носитель "{disc["name"]}" удален успешно!', 'warning')
    return redirect(url_for('discs'))


@app.route('/discs/edit/<int:disc_id>', methods=['POST'])
def edit_disc_route(disc_id):
    """Редактирование носителя."""
    disc = get_disc_by_id(disc_id)
    if disc:
        name = request.form.get('name')
        description = request.form.get('description')
        if name:
            update_disc(disc_id, name, description)
            flash(f'Носитель "{name}" обновлен успешно!', 'success')
    return redirect(url_for('discs'))


# === Вкладка: Коробки (Boxes) ===
@app.route('/boxes')
def boxes():
    """Страница управления коробками."""
    boxes = get_all_boxes()
    return render_template('boxes.html', boxes=boxes)


@app.route('/boxes/add', methods=['POST'])
def add_box_route():
    """Добавление новой коробки."""
    name = request.form.get('name')
    description = request.form.get('description')
    
    if name:
        add_box(name, description)
        flash(f'Коробка "{name}" добавлена успешно!', 'success')
    
    return redirect(url_for('boxes'))


@app.route('/boxes/delete/<int:box_id>', methods=['POST'])
def delete_box_route(box_id):
    """Удаление коробки."""
    box = get_box_by_id(box_id)
    if box:
        delete_box(box_id)
        flash(f'Коробка "{box["name"]}" удалена успешно!', 'warning')
    return redirect(url_for('boxes'))


@app.route('/boxes/edit/<int:box_id>', methods=['POST'])
def edit_box_route(box_id):
    """Редактирование коробки."""
    box = get_box_by_id(box_id)
    if box:
        name = request.form.get('name')
        description = request.form.get('description')
        if name:
            update_box(box_id, name, description)
            flash(f'Коробка "{name}" обновлена успешно!', 'success')
    return redirect(url_for('boxes'))


# === Вкладка: Комплектация ===
@app.route('/kits')
def kits():
    """Страница комплектации (продукты, носители, коробки)."""
    products = get_all_products()
    discs = get_all_discs()
    boxes = get_all_boxes()
    
    # Подсчитать количество компонентов для каждого продукта
    for product in products:
        components = get_product_components(product.doc_id)
        product.components_count = len(components) if components else 0
        
        # Подготовить данные для компонентов
        current_components_discs = []
        current_components_boxes = []
        current_components_qty_disc = {}
        current_components_qty_box = {}
        
        if components:
            for comp in components:
                if comp.get('disc_id'):
                    current_components_discs.append(comp['disc_id'])
                    current_components_qty_disc[comp['disc_id']] = comp.get('disc_quantity', 0)
                if comp.get('box_id'):
                    current_components_boxes.append(comp['box_id'])
                    current_components_qty_box[comp['box_id']] = comp.get('box_quantity', 0)
        
        product.current_components_discs = current_components_discs
        product.current_components_boxes = current_components_boxes
        product.current_components_qty_disc = current_components_qty_disc
        product.current_components_qty_box = current_components_qty_box
    
    return render_template('kits.html', products=products, discs=discs, boxes=boxes)


# === Вкладка: Остатки ===
@app.route('/stock')
def stock():
    """Страница управления остатками."""
    stock_discs = get_all_stock_discs()
    stock_boxes = get_all_stock_boxes()
    discs = get_all_discs()
    boxes = get_all_boxes()
    
    return render_template('stock.html', stock_discs=stock_discs, stock_boxes=stock_boxes,
                          discs=discs, boxes=boxes)


@app.route('/stock/add', methods=['POST'])
def add_stock_route():
    """Добавление остатка."""
    component_type = request.form.get('type')
    component_id = int(request.form.get('component_id'))
    quantity = int(request.form.get('quantity'))
    
    if component_type == 'disc':
        add_stock_disc(component_id, quantity)
        flash('Остаток носителя добавлен успешно!', 'success')
    elif component_type == 'box':
        add_stock_box(component_id, quantity)
        flash('Остаток коробки добавлен успешно!', 'success')
    
    return redirect(url_for('stock'))


@app.route('/stock/adjust', methods=['POST'])
def adjust_stock_route():
    """Корректировка остатка."""
    component_type = request.form.get('type')
    component_id = request.form.get('component_id')
    quantity = request.form.get('quantity')
    
    print(f"[ADJUST] type={component_type}, component_id={component_id}, quantity={quantity}")
    
    if not component_type or not component_id or not quantity:
        print(f"[ERROR] Missing required fields")
        flash('Ошибка: отсутствуют обязательные поля', 'error')
        return redirect(url_for('stock'))
    
    try:
        component_id = int(component_id)
        quantity = int(quantity)
    except ValueError as e:
        print(f"[ERROR] Invalid values: {e}")
        flash('Ошибка: неверный формат данных', 'error')
        return redirect(url_for('stock'))
    
    if component_type == 'disc':
        result = adjust_stock_disc(component_id, quantity)
        if result:
            flash('Остаток носителя обновлен успешно!', 'success')
        else:
            flash('Ошибка: носитель не найден', 'error')
    elif component_type == 'box':
        result = adjust_stock_box(component_id, quantity)
        if result:
            flash('Остаток коробки обновлен успешно!', 'success')
        else:
            flash('Ошибка: коробка не найдена', 'error')
    else:
        print(f"[ERROR] Invalid component type: {component_type}")
        flash('Ошибка: неверный тип компонента', 'error')
    
    return redirect(url_for('stock'))


# === Вкладка: Просмотр остатков ===
@app.route('/view-stock')
def view_stock():
    """Страница просмотра остатков."""
    products = get_all_products()
    stock_discs = get_all_stock_discs()
    stock_boxes = get_all_stock_boxes()
    
    return render_template('view_stock.html', products=products, 
                          stock_discs=stock_discs, stock_boxes=stock_boxes)


# === Списание комплектов ===
@app.route('/dispatch', methods=['POST'])
def dispatch():
    """Списание одного комплекта (с опциональной датой)."""
    product_id = int(request.form.get('product_id'))
    quantity = int(request.form.get('quantity'))
    dispatch_date = request.form.get('dispatch_date') or None

    success, message = dispatch_product(product_id, quantity, dispatch_date)
    flash(message, 'success' if success else 'error')
    return redirect(url_for('view_stock'))


@app.route('/dispatch-batch', methods=['POST'])
def dispatch_batch():
    """Пакетное списание комплектов по текущей дате."""
    items = request.get_json(silent=True)
    if items is None:
        items = []
        for key, value in request.form.items():
            if key.startswith('quantity_') and value:
                product_id = int(key.replace('quantity_', ''))
                items.append({'product_id': product_id, 'quantity': value})

    dispatch_items = []
    for item in items:
        if isinstance(item, dict):
            dispatch_items.append((int(item['product_id']), item['quantity']))
        else:
            dispatch_items.append(item)

    success, messages = dispatch_products_batch(dispatch_items)
    if success:
        for message in messages:
            flash(message, 'success')
    else:
        for message in messages:
            flash(message, 'error')

    if request.is_json or request.content_type == 'application/json':
        return jsonify({'success': success, 'messages': messages})

    return redirect(url_for('view_stock'))


# === Отчеты ===
@app.route('/reports')
def reports():
    """Страница отчетов."""
    from data import get_all_stock_discs, get_all_stock_boxes, get_all_products
    stock_discs = get_all_stock_discs()
    stock_boxes = get_all_stock_boxes()
    products = get_all_products()
    
    return render_template('reports.html', stock_discs=stock_discs, stock_boxes=stock_boxes, products=products)


@app.route('/reports/operations')
def operations_report():
    """Отчет о операциях за период."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    operations = get_operations()
    
    # Преобразуем Document объекты в словари
    operations = [dict(op) for op in operations]
    
    # Добавляем имя продукта для каждой операции
    for op in operations:
        product_id = op.get('product_id', '')
        if product_id:
            from data import get_product_by_id
            product = get_product_by_id(product_id)
            if product:
                op['product_name'] = product['name']
            else:
                op['product_name'] = 'Неизвестно'
        else:
            op['product_name'] = '-'
    
    # Фильтрация по дате если указаны параметры
    if start_date:
        try:
            from datetime import datetime
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            operations = [op for op in operations if datetime.fromisoformat(op.get('date', '').replace(' ', 'T')) >= start_dt]
        except ValueError:
            pass
    
    if end_date:
        try:
            from datetime import datetime
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            operations = [op for op in operations if datetime.fromisoformat(op.get('date', '').replace(' ', 'T')) <= end_dt]
        except ValueError:
            pass
    
    return render_template('operations_report.html', operations=operations, 
                          start_date=start_date, end_date=end_date)


@app.route('/reports/operations/export')
def operations_report_export():
    """Экспорт отчета о операциях в Excel."""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    operations = get_operations()
    
    # Фильтрация по дате если указаны параметры
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            operations = [op for op in operations if datetime.fromisoformat(op.get('date', '').replace(' ', 'T')) >= start_dt]
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            operations = [op for op in operations if datetime.fromisoformat(op.get('date', '').replace(' ', 'T')) <= end_dt]
        except ValueError:
            pass
    
    # Создание книги
    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Заголовок отчета
    ws['A1'] = "ОТЧЕТ О ОПЕРАЦИЯХ"
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = center_alignment
    
    period_str = f"{start_date or 'Начало'} - {end_date or 'Текущее время'}"
    ws['A2'] = f"Период: {period_str}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:E2')
    
    # Заголовки таблицы
    headers = ['Дата и время', 'Тип операции', 'Продукт', 'Количество', 'Детали']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Маппинг типов операций
    operation_type_map = {
        'add_disc': 'Добавлен носитель',
        'add_box': 'Добавлена коробка',
        'adjust_stock': 'Корректировка остатка',
        'dispatch': 'Списание комплекта'
    }
    
    # Данные
    row = 5
    for op in operations:
        date_str = op.get('date', '')
        op_type = operation_type_map.get(op.get('operation_type', ''), op.get('operation_type', ''))
        
        product_id = op.get('product_id', '')
        if product_id:
            product = get_product_by_id(product_id)
            product_name = product['name'] if product else 'Неизвестно'
        else:
            product_name = '-'
        
        quantity = op.get('quantity', '')
        
        # Детали
        details_text = ''
        details = op.get('details', {})
        if isinstance(details, dict):
            components = details.get('components', [])
            if components:
                details_text = '\n'.join([
                    f"{c.get('type', '').upper()}: {c.get('name', '')} ({c.get('quantity', c.get('total_quantity', ''))} шт.)"
                    for c in components
                ])
        
        ws[f'A{row}'] = date_str
        ws[f'B{row}'] = op_type
        ws[f'C{row}'] = product_name
        ws[f'D{row}'] = quantity
        ws[f'E{row}'] = details_text
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].border = border
            if col == 'D':
                ws[f'{col}{row}'].alignment = center_alignment
            else:
                ws[f'{col}{row}'].alignment = left_alignment
        
        ws.row_dimensions[row].height = max(20, len(details_text.split('\n')) * 15) if details_text else 20
        row += 1
    
    # Установка ширины колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40
    
    # Сохранение в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'операции_{start_date or "all"}_{end_date or "now"}.xlsx'
    )


@app.route('/reports/stock-levels')
def stock_levels_report():
    """Отчет об уровнях остатков."""
    products = get_all_products()
    stock_discs = get_all_stock_discs()
    stock_boxes = get_all_stock_boxes()
    
    return jsonify({
        'products': products,
        'stock_discs': stock_discs,
        'stock_boxes': stock_boxes
    })


@app.route('/reports/stock-levels/export')
def stock_levels_export():
    """Экспорт отчета об уровнях остатков в Excel."""
    products = get_all_products()
    stock_discs = get_all_stock_discs()
    stock_boxes = get_all_stock_boxes()
    
    # Создание книги
    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Заголовок отчета
    ws['A1'] = "ОТЧЕТ ПО ОСТАТКАМ"
    ws['A1'].font = title_font
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = center_alignment
    
    ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:B2')
    
    # Продукты
    row = 4
    ws[f'A{row}'] = "ПРОДУКТЫ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'B{row}'] = ""
    ws[f'B{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Продукт"
    ws[f'B{row}'] = "Доступно (комплектов)"
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].alignment = center_alignment
        ws[f'{col}{row}'].border = border
    
    row += 1
    for product in products:
        product_id = product.doc_id
        name = product['name']
        components = get_product_components(product_id)
        
        # Определить минимальное доступное количество комплектов
        min_quantity = float('inf')
        for comp in components:
            disc_qty = get_stock_disc_quantity(comp.get('disc_id', 0)) if comp.get('disc_id') else float('inf')
            box_qty = get_stock_box_quantity(comp.get('box_id', 0)) if comp.get('box_id') else float('inf')
            
            if comp.get('disc_id') and comp.get('disc_quantity', 0) > 0:
                disc_qty //= comp['disc_quantity']
            if comp.get('box_id') and comp.get('box_quantity', 0) > 0:
                box_qty //= comp['box_quantity']
            
            min_quantity = min(min_quantity, disc_qty, box_qty)
        
        if min_quantity == float('inf'):
            min_quantity = 0
        
        ws[f'A{row}'] = name
        ws[f'B{row}'] = int(min_quantity)
        ws[f'A{row}'].alignment = left_alignment
        ws[f'B{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Носители
    row += 1
    ws[f'A{row}'] = "НОСИТЕЛИ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'B{row}'] = ""
    ws[f'B{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Носитель"
    ws[f'B{row}'] = "Остаток"
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].alignment = center_alignment
        ws[f'{col}{row}'].border = border
    
    row += 1
    for disc in stock_discs:
        ws[f'A{row}'] = disc['disc_name']
        ws[f'B{row}'] = disc['quantity']
        ws[f'A{row}'].alignment = left_alignment
        ws[f'B{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Коробки
    row += 1
    ws[f'A{row}'] = "КОРОБКИ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'B{row}'] = ""
    ws[f'B{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Коробка"
    ws[f'B{row}'] = "Остаток"
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].fill = header_fill
        ws[f'{col}{row}'].alignment = center_alignment
        ws[f'{col}{row}'].border = border
    
    row += 1
    for box in stock_boxes:
        ws[f'A{row}'] = box['box_name']
        ws[f'B{row}'] = box['quantity']
        ws[f'A{row}'].alignment = left_alignment
        ws[f'B{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Установка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    
    # Сохранение в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'остатки_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# === Экспорт остатков со страницы Списание ===
@app.route('/export-stock-to-excel', methods=['POST'])
def export_stock_to_excel():
    """Экспорт остатков в Excel."""
    data = request.get_json()
    
    # Создание книги
    wb = Workbook()
    ws = wb.active
    ws.title = "Остатки"
    
    # Стили
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Заголовок отчета
    ws['A1'] = "ОСТАТКИ КОМПЛЕКТОВ"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = center_alignment
    
    ws['A2'] = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:D2')
    
    # Продукты
    row = 4
    ws[f'A{row}'] = "ПРОДУКТЫ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    for col in ['B', 'C', 'D']:
        ws[f'{col}{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    headers = ['Продукт', 'Доступно', 'Носителей', 'Коробок']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    row += 1
    for product in data.get('products', []):
        ws[f'A{row}'] = product['name']
        ws[f'B{row}'] = product['available']
        ws[f'C{row}'] = product['discs']
        ws[f'D{row}'] = product['boxes']
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = border
            if col == 'A':
                ws[f'{col}{row}'].alignment = left_alignment
            else:
                ws[f'{col}{row}'].alignment = center_alignment
        row += 1
    
    # Носители
    row += 1
    ws[f'A{row}'] = "НОСИТЕЛИ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'B{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    for col_num, header in enumerate(['Носитель', 'Остаток'], 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    row += 1
    for disc in data.get('discs', []):
        ws[f'A{row}'] = disc['name']
        ws[f'B{row}'] = disc['quantity']
        ws[f'A{row}'].alignment = left_alignment
        ws[f'B{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Коробки
    row += 1
    ws[f'A{row}'] = "КОРОБКИ"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = section_fill
    ws[f'B{row}'].fill = section_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    for col_num, header in enumerate(['Коробка', 'Остаток'], 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    row += 1
    for box in data.get('boxes', []):
        ws[f'A{row}'] = box['name']
        ws[f'B{row}'] = box['quantity']
        ws[f'A{row}'].alignment = left_alignment
        ws[f'B{row}'].alignment = center_alignment
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1
    
    # Установка ширины колонок
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    # Сохранение в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'остатки_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


# === История операций ===
@app.route('/operations')
def operations():
    """Страница истории операций."""
    from data import get_operations, get_product_by_id
    operations = get_operations()
    # Преобразуем Document объекты в словари
    operations = [dict(op) for op in operations]
    # Добавляем имя продукта и обработку типа операции для каждой операции
    for op in operations:
        product_id = op.get('product_id', '')
        if product_id:
            product = get_product_by_id(product_id)
            if product:
                op['product_name'] = product['name']
            else:
                op['product_name'] = 'Неизвестно'
        else:
            op['product_name'] = '-'
        
        # Гарантируем что operation_type существует
        if 'operation_type' not in op:
            op['operation_type'] = ''
    return render_template('operations.html', operations=operations)


@app.route('/api/operations')
def api_operations():
    """API для получения всех операций."""
    operations = get_operations()
    # Преобразуем Document объекты в словари для JSON
    result = []
    for op in operations:
        op_dict = dict(op)
        # Добавляем имя продукта для удобства
        product_id = op.get('product_id', '')
        if product_id:
            product = get_product_by_id(product_id)
            if product:
                op_dict['product_name'] = product['name']
        result.append(op_dict)
    return jsonify(result)


# === API endpoints ===
@app.route('/api/products')
def api_products():
    """API для получения всех продуктов."""
    products = get_all_products()
    return jsonify(products)


@app.route('/api/discs')
def api_discs():
    """API для получения всех носителей."""
    discs = get_all_discs()
    return jsonify(discs)


@app.route('/api/boxes')
def api_boxes():
    """API для получения всех коробок."""
    boxes = get_all_boxes()
    return jsonify(boxes)


@app.route('/api/stock')
def api_stock():
    """API для получения всех остатков."""
    stock_discs = get_all_stock_discs()
    stock_boxes = get_all_stock_boxes()
    return jsonify({
        'discs': stock_discs,
        'boxes': stock_boxes
    })


@app.route('/api/stock/disc/<int:disc_id>')
def api_stock_disc(disc_id):
    """API для получения остатка диска."""
    from data import get_stock_disc_quantity
    quantity = get_stock_disc_quantity(disc_id)
    return jsonify({'quantity': quantity})


@app.route('/api/stock/box/<int:box_id>')
def api_stock_box(box_id):
    """API для получения остатка коробки."""
    from data import get_stock_box_quantity
    quantity = get_stock_box_quantity(box_id)
    return jsonify({'quantity': quantity})


if __name__ == '__main__':
    # Запуск веб-сервера
    app.run(debug=True, host='0.0.0.0', port=5000)
