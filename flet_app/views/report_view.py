import flet as ft
from data import (
    get_all_products, get_product_components,
    get_stock_disc_quantity, get_stock_box_quantity,
    get_all_stock_discs, get_all_stock_boxes,
)
from datetime import datetime


class ReportDialog:
    def __init__(self, page):
        self.page = page
        self.dlg = None

    def open(self):
        products = get_all_products()
        discs = get_all_stock_discs()
        boxes = get_all_stock_boxes()

        rows = []

        for p in products:
            components = get_product_components(p.doc_id)
            min_qty = float('inf')
            for c in components:
                if c.get('disc_id') and c.get('disc_quantity', 0) > 0:
                    dq = get_stock_disc_quantity(c['disc_id']) // c['disc_quantity']
                    min_qty = min(min_qty, dq)
                if c.get('box_id') and c.get('box_quantity', 0) > 0:
                    bq = get_stock_box_quantity(c['box_id']) // c['box_quantity']
                    min_qty = min(min_qty, bq)
            available = 0 if min_qty == float('inf') else int(min_qty)
            rows.append(('Продукт', p['name'], available))

        for d in discs:
            rows.append(('Носитель', d['disc_name'], d['quantity']))

        for b in boxes:
            rows.append(('Коробка', b['box_name'], b['quantity']))

        table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Тип")),
                ft.DataColumn(ft.Text("Наименование")),
                ft.DataColumn(ft.Text("Доступно / Остаток")),
            ],
            rows=[
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(t, weight=ft.FontWeight.W_500 if t == 'Продукт' else ft.FontWeight.NORMAL,
                                        color=ft.Colors.BLUE if t == 'Продукт' else ft.Colors.GREEN if t == 'Носитель' else ft.Colors.ORANGE)),
                    ft.DataCell(ft.Text(n)),
                    ft.DataCell(ft.Text(str(q))),
                ])
                for t, n, q in rows
            ] if rows else [
                ft.DataRow(cells=[ft.DataCell(ft.Text("Нет данных")) for _ in range(3)])
            ],
            border=ft.border.Border(
                left=ft.BorderSide(1, ft.Colors.OUTLINE),
                top=ft.BorderSide(1, ft.Colors.OUTLINE),
                right=ft.BorderSide(1, ft.Colors.OUTLINE),
                bottom=ft.BorderSide(1, ft.Colors.OUTLINE),
            ),
            border_radius=8,
            vertical_lines=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
        )

        self.dlg = ft.AlertDialog(
            title=ft.Text("Отчет по остаткам"),
            content=ft.Column([
                ft.Text(f"Дата отчета: {datetime.now().strftime('%d.%m.%Y')}", size=12, color=ft.Colors.GREY),
                ft.Divider(height=5),
                ft.Column([table], expand=True, scroll=ft.ScrollMode.AUTO),
            ], width=700, height=500),
            actions=[
                ft.FilledButton("Экспорт в Excel", icon=ft.Icons.TABLE_CHART, on_click=lambda e: self._export_excel(rows)),
                ft.TextButton("Закрыть", on_click=lambda e: self.close()),
            ],
        )
        self.page.show_dialog(self.dlg)
        self.page.update()

    def close(self):
        if self.dlg:
            self.page.pop_dialog()

    def _export_excel(self, rows):
        async def do_export():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            except ImportError:
                dlg = ft.AlertDialog(
                    title=ft.Text("Ошибка"),
                    content=ft.Text("Установите openpyxl: pip install openpyxl"),
                    actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())],
                )
                self.page.show_dialog(dlg)
                return

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Остатки комплектов"

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'),
                )

                headers = ['Тип', 'Наименование', 'Доступно / Остаток']
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                for i, (t, n, q) in enumerate(rows, 2):
                    ws.cell(row=i, column=1, value=t).border = thin_border
                    ws.cell(row=i, column=2, value=n).border = thin_border
                    ws.cell(row=i, column=3, value=q).border = thin_border

                ws.column_dimensions['A'].width = 14
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 22

                ws['E1'] = "Дата отчета:"
                ws['F1'] = datetime.now().strftime('%d.%m.%Y')

                default_name = f"Отчет_по_остаткам_ИК_{datetime.now().strftime('%Y%m%d')}.xlsx"

                file_picker = ft.FilePicker()
                self.page.overlay.append(file_picker)
                self.page.update()
                result = await file_picker.save_file(
                    file_name=default_name,
                    allowed_extensions=["xlsx"],
                )
                if result:
                    wb.save(result)
                    dlg = ft.AlertDialog(
                        title=ft.Text("Excel"),
                        content=ft.Text(f"Отчет сохранен: {result}"),
                        actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())],
                    )
                    self.page.show_dialog(dlg)

            except Exception as ex:
                dlg = ft.AlertDialog(
                    title=ft.Text("Ошибка"),
                    content=ft.Text(f"Ошибка сохранения: {ex}"),
                    actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())],
                )
                self.page.show_dialog(dlg)

        self.page.run_task(do_export)
