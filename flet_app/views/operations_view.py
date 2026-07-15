import flet as ft
from data import get_operations, get_product_by_id, get_all_products
from datetime import datetime, date


OPERATION_LABELS = {
    'dispatch': 'Списание',
    'add_disc': 'Приход дисков',
    'add_box': 'Приход коробок',
    'adjust_stock': 'Корректировка',
    'write_off': 'Списание по браку',
    'add_kit': 'Добавление комплекта',
}

OPERATION_COLORS = {
    'dispatch': ft.Colors.RED,
    'add_disc': ft.Colors.GREEN,
    'add_box': ft.Colors.GREEN,
    'adjust_stock': ft.Colors.ORANGE,
    'write_off': ft.Colors.PURPLE,
    'add_kit': ft.Colors.BLUE,
}


def _format_op_date(d):
    if len(d) >= 10:
        return d[8:10] + "." + d[5:7] + "." + d[:4] + d[10:16]
    return d

class OperationsView:
    def __init__(self, app):
        self.app = app
        self.page = app.page
        self.table = None
        self.filter_dd = None
        self.start_date = None
        self.end_date = None
        self.product_dd = None

    @staticmethod
    def _parse_disp_date(val):
        v = val.strip() if val else ""
        if not v:
            return ""
        parts = v.split(".")
        if len(parts) == 3:
            return f"{parts[2]}-{parts[1]}-{parts[0]}"
        return v

    def refresh(self):
        pass

    def _build_date_field(self, label):
        field = ft.TextField(label=label, value="", width=180)

        def on_date_selected(ev):
            val = ev.data
            if isinstance(val, datetime):
                field.value = val.astimezone().strftime("%d.%m.%Y")
            elif isinstance(val, date):
                field.value = val.strftime("%d.%m.%Y")
            elif isinstance(val, str) and len(val) >= 10:
                field.value = val[8:10] + "." + val[5:7] + "." + val[:4]
            else:
                try:
                    val = ev.control.value
                    if isinstance(val, datetime):
                        field.value = val.astimezone().strftime("%d.%m.%Y")
                    elif isinstance(val, date):
                        field.value = val.strftime("%d.%m.%Y")
                    elif isinstance(val, str) and len(val) >= 10:
                        field.value = val[8:10] + "." + val[5:7] + "." + val[:4]
                except Exception:
                    pass
            self.page.update()

        picker = ft.DatePicker(
            on_change=on_date_selected,
            locale=ft.Locale("ru"),
        )

        def open_picker(e):
            self.page.show_dialog(picker)

        btn = ft.IconButton(
            ft.Icons.CALENDAR_MONTH,
            tooltip="Выбрать дату",
            on_click=open_picker,
        )
        return ft.Row([field, btn], spacing=5)

    def build(self):
        header = ft.Text("История операций", size=24, weight=ft.FontWeight.BOLD)

        start_date_row = self._build_date_field("Дата с (ДД.ММ.ГГГГ)")
        self.start_date = start_date_row.controls[0]
        end_date_row = self._build_date_field("Дата по (ДД.ММ.ГГГГ)")
        self.end_date = end_date_row.controls[0]

        self.filter_dd = ft.Dropdown(
            label="Фильтр по типу",
            options=[
                ft.dropdown.Option("all", "Все операции"),
                ft.dropdown.Option("dispatch", "Списание"),
                ft.dropdown.Option("add_disc", "Приход дисков"),
                ft.dropdown.Option("add_box", "Приход коробок"),
                ft.dropdown.Option("adjust_stock", "Корректировка"),
                ft.dropdown.Option("write_off", "Списание по браку"),
            ],
            value="all",
            width=200,
        )
        self.filter_dd.on_change = self.on_filter_change

        product_names = ["Все"] + [p['name'] for p in get_all_products()]
        self.product_dd = ft.Dropdown(
            label="Фильтр по продукту",
            options=[ft.dropdown.Option(n, n) for n in product_names],
            value="Все",
            width=250,
        )
        self.product_dd.on_change = self.on_filter_change

        filter_row1 = ft.Row([
            start_date_row,
            end_date_row,
            ft.FilledButton("Применить", icon=ft.Icons.SEARCH, on_click=self.on_filter_change),
            ft.FilledButton("Сбросить", icon=ft.Icons.CLEAR, on_click=self.reset_filter),
        ], spacing=10)

        filter_row2 = ft.Row([
            self.filter_dd,
            self.product_dd,
            ft.FilledButton("Экспорт в Excel", icon=ft.Icons.TABLE_CHART, on_click=self.export_excel),
        ], spacing=10)

        self.table = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Дата")),
                ft.DataColumn(ft.Text("Тип")),
                ft.DataColumn(ft.Text("Продукт")),
                ft.DataColumn(ft.Text("Количество")),
                ft.DataColumn(ft.Text("Детали")),
            ],
            rows=[],
            border=ft.border.Border(
                left=ft.BorderSide(1, ft.Colors.OUTLINE),
                top=ft.BorderSide(1, ft.Colors.OUTLINE),
                right=ft.BorderSide(1, ft.Colors.OUTLINE),
                bottom=ft.BorderSide(1, ft.Colors.OUTLINE),
            ),
            border_radius=8,
            vertical_lines=ft.BorderSide(1, ft.Colors.OUTLINE_VARIANT),
        )

        self.load_operations()

        return ft.Column([
            header,
            ft.Divider(),
            filter_row1,
            filter_row2,
            ft.Divider(),
            ft.Column([self.table], expand=True, scroll=ft.ScrollMode.AUTO),
        ], expand=True, scroll=ft.ScrollMode.AUTO)

    def on_filter_change(self, e=None):
        self.load_operations()

    def reset_filter(self, e=None):
        if self.start_date:
            self.start_date.value = ""
        if self.end_date:
            self.end_date.value = ""
        if self.filter_dd:
            self.filter_dd.value = "all"
        if self.product_dd:
            self.product_dd.value = "Все"
        self.load_operations()

    def load_operations(self):
        filter_type = self.filter_dd.value if self.filter_dd else "all"
        start = self._parse_disp_date(self.start_date.value.strip() if self.start_date else "")
        end = self._parse_disp_date(self.end_date.value.strip() if self.end_date else "")
        filter_product = self.product_dd.value if self.product_dd else "Все"

        ops = get_operations()

        if filter_type and filter_type != "all":
            ops = [o for o in ops if o.get('operation_type') == filter_type]

        if start:
            ops = [o for o in ops if o.get('date', '')[:10] >= start]

        if end:
            ops = [o for o in ops if o.get('date', '')[:10] <= end]

        if filter_product and filter_product != "Все":
            filtered = []
            for op in ops:
                pid = op.get('product_id')
                if pid:
                    product = get_product_by_id(pid)
                    if product and product['name'] == filter_product:
                        filtered.append(op)
                elif filter_product == "-":
                    filtered.append(op)
            ops = filtered

        rows = []
        for op in ops:
            op_type = op.get('operation_type', '')
            label = OPERATION_LABELS.get(op_type, op_type)
            color = OPERATION_COLORS.get(op_type, ft.Colors.GREY)
            details = op.get('details', {})
            components = details.get('components', [])
            detail_text = '; '.join(
                f"{c.get('name', '')} ({c.get('total_quantity', c.get('quantity', ''))} шт.)"
                for c in components
            ) if components else ''

            pid = op.get('product_id')
            product_name = str(pid) if pid else '—'
            if pid:
                product = get_product_by_id(pid)
                if product:
                    product_name = product['name']

            rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(_format_op_date(op.get('date', '')))),
                    ft.DataCell(ft.Text(label, color=color, weight=ft.FontWeight.W_500)),
                    ft.DataCell(ft.Text(product_name)),
                    ft.DataCell(ft.Text(str(op.get('quantity', '')))),
                    ft.DataCell(ft.Text(detail_text, selectable=True)),
                ])
            )

        self.table.rows = rows if rows else [
            ft.DataRow(cells=[ft.DataCell(ft.Text("Нет операций")) for _ in range(5)])
        ]
        if self.page:
            self.page.update()

    def _get_filtered_ops(self):
        filter_type = self.filter_dd.value if self.filter_dd else "all"
        start = self._parse_disp_date(self.start_date.value.strip() if self.start_date else "")
        end = self._parse_disp_date(self.end_date.value.strip() if self.end_date else "")
        filter_product = self.product_dd.value if self.product_dd else "Все"

        ops = get_operations()

        if filter_type and filter_type != "all":
            ops = [o for o in ops if o.get('operation_type') == filter_type]

        if start:
            ops = [o for o in ops if o.get('date', '')[:10] >= start]

        if end:
            ops = [o for o in ops if o.get('date', '')[:10] <= end]

        if filter_product and filter_product != "Все":
            filtered = []
            for op in ops:
                pid = op.get('product_id')
                if pid:
                    product = get_product_by_id(pid)
                    if product and product['name'] == filter_product:
                        filtered.append(op)
                elif filter_product == "-":
                    filtered.append(op)
            ops = filtered
        return ops

    def export_excel(self, e):
        file_picker = ft.FilePicker()
        self.page.overlay.append(file_picker)

        async def do_export(fp):
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            except ImportError:
                dlg = ft.AlertDialog(
                    title=ft.Text("Ошибка"),
                    content=ft.Text("Установите openpyxl: pip install openpyxl"),
                    actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())],
                )
                self.page.show_dialog(dlg)
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Операции"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )

            headers = ['ID', 'Дата и время', 'Тип операции', 'Продукт', 'Количество', 'Детали']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            ops = self._get_filtered_ops()
            for row_idx, op in enumerate(ops, 2):
                pid = op.get('product_id')
                product_name = str(pid) if pid else '—'
                if pid:
                    product = get_product_by_id(pid)
                    if product:
                        product_name = product['name']

                details = op.get('details', {})
                components = details.get('components', [])
                detail_text = '; '.join(
                    f"{c.get('name', '')} ({c.get('total_quantity', c.get('quantity', ''))} шт.)"
                    for c in components
                ) if components else ''
                reason = details.get('reason', '')
                if reason:
                    detail_text += f" | Причина: {reason}" if detail_text else f"Причина: {reason}"

                op_type = OPERATION_LABELS.get(op.get('operation_type', ''), op.get('operation_type', ''))

                ws.cell(row=row_idx, column=1, value=op.doc_id).border = thin_border
                ws.cell(row=row_idx, column=2, value=op.get('date', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value=op_type).border = thin_border
                ws.cell(row=row_idx, column=4, value=product_name).border = thin_border
                ws.cell(row=row_idx, column=5, value=op.get('quantity', '')).border = thin_border
                ws.cell(row=row_idx, column=6, value=detail_text).border = thin_border

            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 22
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 45

            default_name = f"Операции_{datetime.now().strftime('%Y%m%d')}.xlsx"
            result = await fp.save_file(
                file_name=default_name,
                allowed_extensions=["xlsx"],
            )
            if result:
                try:
                    wb.save(result)
                    dlg = ft.AlertDialog(
                        title=ft.Text("Excel"),
                        content=ft.Text(f"Отчет сохранен: {result}"),
                        actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())],
                    )
                    self.page.show_dialog(dlg)
                except Exception as ex:
                    dlg = ft.AlertDialog(
                        title=ft.Text("Ошибка"),
                        content=ft.Text(f"Ошибка сохранения: {ex}"),
                        actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())],
                    )
                    self.page.show_dialog(dlg)

        self.page.run_task(do_export, file_picker)
