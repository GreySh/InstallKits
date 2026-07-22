"""
Формирование красивых отчётов Excel (в стиле веб-версии / Toga).
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox

from data import (
    get_all_products, get_product_available_quantity,
    get_all_stock_discs, get_all_stock_boxes,
)


def build_web_report(title, subtitle, sections, widths, wrap=False):
    """Построить книгу в формате веб-версии / Toga.

    sections — список (заголовок_секции, заголовки, строки, left_columns),
    где left_columns — множество 0-based индексов столбцов с выравниванием влево.
    Возвращает (wb, ws).
    """
    HEADER_FILL = PatternFill("solid", fgColor="366092")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    TITLE_FONT = Font(bold=True, size=14)
    SUBTITLE_FONT = Font(italic=True, size=10)
    SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
    SECTION_FILL = PatternFill("solid", fgColor="4472C4")
    BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    CENTER = Alignment(horizontal='center', vertical='center')
    LEFT = Alignment(horizontal='left', vertical='center')
    CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_c = CENTER_WRAP if wrap else CENTER
    align_l = LEFT_WRAP if wrap else LEFT

    wb = Workbook()
    ws = wb.active
    ws.title = title
    ncols = max((len(h) for _, h, _, _ in sections), default=1)
    last_col = get_column_letter(ncols)
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].alignment = align_c
    ws['A2'] = subtitle
    ws['A2'].font = SUBTITLE_FONT
    ws.merge_cells(f'A2:{last_col}2')
    row = 4
    first_data_row = None
    for sec_title, headers, rows, left_columns in sections:
        if left_columns is None:
            left_columns = {0}
        ws[f'A{row}'] = sec_title
        ws[f'A{row}'].font = SECTION_FONT
        ws[f'A{row}'].fill = SECTION_FILL
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        row += 1
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci)
            cell.value = h
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = align_c
            cell.border = BORDER
        row += 1
        if first_data_row is None:
            first_data_row = row
        for ri, vals in enumerate(rows):
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci)
                cell.value = v
                cell.alignment = align_l if (ci - 1) in left_columns else align_c
                cell.border = BORDER
                if ri % 2 == 1:
                    cell.fill = PatternFill("solid", fgColor="F2F6FA")
            row += 1
        row += 1
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    if first_data_row is not None:
        ws.freeze_panes = ws.cell(row=first_data_row, column=1)
    return wb, ws


def build_stock_report():
    """Сформировать отчёт по остаткам (продукты, носители, коробки)."""
    product_rows = [[p["name"], get_product_available_quantity(p.doc_id)]
                    for p in get_all_products()]
    disc_rows = [[d["disc_name"], d["quantity"]] for d in get_all_stock_discs()]
    box_rows = [[b["box_name"], b["quantity"]] for b in get_all_stock_boxes()]
    return build_web_report(
        "ОТЧЕТ ПО ОСТАТКАМ",
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}",
        [
            ("ПРОДУКТЫ", ["Продукт", "Доступно (комплектов)"], product_rows, {0}),
            ("НОСИТЕЛИ", ["Носитель", "Остаток"], disc_rows, {0}),
            ("КОРОБКИ", ["Коробка", "Остаток"], box_rows, {0}),
        ],
        widths={'A': 35, 'B': 25},
    )


def export_stock_report(parent=None):
    """Сформировать и сохранить отчёт по остаткам через диалог сохранения."""
    wb, _ = build_stock_report()
    suggested = f"Отчет по остаткам ИК_{datetime.now().strftime('%Y%m%d')}.xlsx"

    try:
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=suggested,
            parent=parent,
            title="Сохранить отчёт по остаткам",
        )
    except Exception:
        return

    if not file_path:
        return

    try:
        wb.save(file_path)
        messagebox.showinfo("Успех", f"Файл сохранён:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")
