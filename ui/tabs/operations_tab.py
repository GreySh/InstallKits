"""
Вкладка истории операций с фильтрацией по дате, типу и продукту, экспортом в Excel.
"""

import customtkinter as ctk
from tkinter import ttk, messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os

from data import get_operations, get_product_by_id, get_all_products
from ui.widgets.date_picker import DatePicker


OPERATION_TYPE_MAP = {
    'add_disc': 'Добавлен носитель',
    'add_box': 'Добавлена коробка',
    'adjust_stock': 'Корректировка остатка',
    'dispatch': 'Списание комплекта',
    'write_off': 'Списание по браку',
}


class OperationsTab(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # Заголовок
        ctk.CTkLabel(self, text="История операций", font=("Arial", 16, "bold")).grid(
            row=0, column=0, padx=10, pady=10, sticky="w"
        )

        # Фильтры
        filter_frame = ctk.CTkFrame(self)
        filter_frame.grid(row=1, column=0, padx=10, pady=5, sticky="ew")

        # Строка 1: даты
        row1 = ctk.CTkFrame(filter_frame, fg_color="transparent")
        row1.pack(fill="x", padx=5, pady=(5, 2))

        ctk.CTkLabel(row1, text="С:").pack(side="left", padx=(10, 2))
        self.start_date = DatePicker(row1)
        self.start_date.pack(side="left", padx=2)

        ctk.CTkLabel(row1, text="По:").pack(side="left", padx=(10, 2))
        self.end_date = DatePicker(row1)
        self.end_date.pack(side="left", padx=2)

        ctk.CTkButton(row1, text="Показать", command=self.load_operations).pack(side="left", padx=10)
        ctk.CTkButton(row1, text="Сбросить", command=self.reset_filter).pack(side="left", padx=5)

        # Строка 2: тип операции + продукт
        row2 = ctk.CTkFrame(filter_frame, fg_color="transparent")
        row2.pack(fill="x", padx=5, pady=(2, 5))

        ctk.CTkLabel(row2, text="Тип операции:").pack(side="left", padx=(10, 2))
        self.type_var = ctk.StringVar(value="Все")
        type_values = ["Все"] + list(OPERATION_TYPE_MAP.values())
        self.type_combo = ctk.CTkComboBox(row2, values=type_values, variable=self.type_var,
                                           width=180, command=lambda _: self.load_operations())
        self.type_combo.pack(side="left", padx=2)

        ctk.CTkLabel(row2, text="Продукт:").pack(side="left", padx=(10, 2))
        self.product_var = ctk.StringVar(value="Все")
        product_names = ["Все"] + [p['name'] for p in get_all_products()]
        self.product_combo = ctk.CTkComboBox(row2, values=product_names, variable=self.product_var,
                                              width=180, command=lambda _: self.load_operations())
        self.product_combo.pack(side="left", padx=2)

        # Таблица операций
        self.ops_tree = ttk.Treeview(
            self,
            columns=("date", "type", "product", "quantity", "details"),
            show="tree headings",
        )
        self.ops_tree.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

        self.ops_tree.heading("#0", text="ID", command=lambda: self.sort_by("#0"))
        self.ops_tree.heading("date", text="Дата и время", command=lambda: self.sort_by("date"))
        self.ops_tree.heading("type", text="Тип операции", command=lambda: self.sort_by("type"))
        self.ops_tree.heading("product", text="Продукт", command=lambda: self.sort_by("product"))
        self.ops_tree.heading("quantity", text="Количество", command=lambda: self.sort_by("quantity"))
        self.ops_tree.heading("details", text="Детали", command=lambda: self.sort_by("details"))

        self.ops_tree.column("#0", width=40, anchor="center")
        self.ops_tree.column("date", width=140, anchor="w")
        self.ops_tree.column("type", width=160, anchor="w")
        self.ops_tree.column("product", width=120, anchor="w")
        self.ops_tree.column("quantity", width=70, anchor="center")
        self.ops_tree.column("details", width=250, anchor="w")

        # Кнопки внизу
        btn_frame = ctk.CTkFrame(self)
        btn_frame.grid(row=3, column=0, padx=10, pady=5, sticky="e")

        ctk.CTkButton(btn_frame, text="Экспорт в Excel", command=self.export_excel).pack(side="right", padx=5)
        ctk.CTkButton(btn_frame, text="Обновить", command=self.load_operations).pack(side="right", padx=5)

        # Состояние сортировки
        self.sort_column = None
        self.sort_reverse = False

        # Загрузить данные
        self.load_operations()

    def load_operations(self):
        """Загрузить операции с учётом всех фильтров."""
        for item in self.ops_tree.get_children():
            self.ops_tree.delete(item)

        start = self.start_date.get().strip()
        end = self.end_date.get().strip()
        filter_type = self.type_var.get()
        filter_product = self.product_var.get()

        # Обратное маппание: русское название -> ключ
        reverse_type_map = {v: k for k, v in OPERATION_TYPE_MAP.items()}
        filter_type_key = reverse_type_map.get(filter_type) if filter_type != "Все" else None

        operations = get_operations()

        # Фильтр по дате
        if start:
            try:
                start_dt = datetime.strptime(start, '%d.%m.%Y').strftime('%Y-%m-%d')
                operations = [op for op in operations if op.get('date', '')[:10] >= start_dt]
            except ValueError:
                pass

        if end:
            try:
                end_dt = datetime.strptime(end, '%d.%m.%Y').strftime('%Y-%m-%d')
                operations = [op for op in operations if op.get('date', '')[:10] <= end_dt]
            except ValueError:
                pass

        # Фильтр по типу операции
        if filter_type_key:
            operations = [op for op in operations if op.get('operation_type') == filter_type_key]

        # Фильтр по продукту
        if filter_product != "Все":
            filtered = []
            for op in operations:
                pid = op.get('product_id')
                if pid:
                    product = get_product_by_id(pid)
                    if product and product['name'] == filter_product:
                        filtered.append(op)
                elif filter_product == "-":
                    filtered.append(op)
            operations = filtered

        for op in operations:
            op_id = op.doc_id
            date = op.get('date', '')
            op_type = OPERATION_TYPE_MAP.get(op.get('operation_type', ''), op.get('operation_type', ''))

            product_id = op.get('product_id')
            product_name = '-'
            if product_id:
                product = get_product_by_id(product_id)
                product_name = product['name'] if product else 'Неизвестно'

            quantity = op.get('quantity', '')

            details = op.get('details', {})
            details_text = ''
            if isinstance(details, dict):
                components = details.get('components', [])
                if components:
                    parts = []
                    for c in components:
                        qty = c.get('quantity', c.get('total_quantity', ''))
                        parts.append(f"{c.get('type', '').upper()}: {c.get('name', '')} ({qty} шт.)")
                    details_text = '; '.join(parts)
                reason = details.get('reason', '')
                if reason:
                    details_text += f" | Причина: {reason}" if details_text else f"Причина: {reason}"

            self.ops_tree.insert(
                "", "end",
                text=str(op_id),
                values=(date, op_type, product_name, quantity, details_text),
            )

    def reset_filter(self):
        """Сбросить все фильтры."""
        self.start_date.delete(0, 'end')
        self.end_date.delete(0, 'end')
        self.type_var.set("Все")
        self.product_var.set("Все")
        self.load_operations()

    def sort_by(self, column):
        """Сортировка по колонке при клике на заголовок."""
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False

        items = []
        for item_id in self.ops_tree.get_children():
            item = self.ops_tree.item(item_id)
            text = item['text']
            values = item['values']
            items.append((text, values))

        # Определяем ключ сортировки
        numeric_cols = {"#0", "quantity"}
        date_cols = {"date"}

        if column in numeric_cols:
            if column == "#0":
                items.sort(key=lambda x: int(x[0]), reverse=self.sort_reverse)
            else:
                items.sort(key=lambda x: int(x[1][4]) if x[1][4] else 0, reverse=self.sort_reverse)
        elif column in date_cols:
            items.sort(key=lambda x: x[1][0] if x[1][0] else "", reverse=self.sort_reverse)
        else:
            col_index = {"type": 1, "product": 2, "details": 3}.get(column, 0)
            items.sort(key=lambda x: x[1][col_index] if x[1][col_index] else "", reverse=self.sort_reverse)

        for item_id in self.ops_tree.get_children():
            self.ops_tree.delete(item_id)

        for text, values in items:
            self.ops_tree.insert("", "end", text=text, values=values)

    def export_excel(self):
        """Экспортировать операции в Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Операции"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        headers = ['ID', 'Дата и время', 'Тип операции', 'Продукт', 'Количество', 'Детали']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        items = self.ops_tree.get_children()
        for row_idx, item_id in enumerate(items, 2):
            item = self.ops_tree.item(item_id)
            values = item['values']
            # values = [date, type, product, quantity, details]
            ws.cell(row=row_idx, column=1, value=int(item['text']))
            ws.cell(row=row_idx, column=2, value=values[0])
            ws.cell(row=row_idx, column=3, value=values[1])
            ws.cell(row=row_idx, column=4, value=values[2])
            ws.cell(row=row_idx, column=5, value=values[3])
            ws.cell(row=row_idx, column=6, value=values[4])

            for col in range(1, 7):
                ws.cell(row=row_idx, column=col).border = thin_border

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 45

        default_filename = f"Операции_{datetime.now().strftime('%Y%m%d')}.xlsx"
        file_path = ctk.filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename,
        )

        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Успех", f"Отчет сохранен в {file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {str(e)}")
