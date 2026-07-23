"""
Окно просмотра зафиксированных остатков на указанную дату.
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date
from data import get_all_fixations
from ui.dialogs.base_dialog import BaseDialog
from ui.widgets.date_picker import DatePicker
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


DATE_FMT = "%d.%m.%Y"
DATE_DB  = "%Y-%m-%d"


class StockViewWindow(BaseDialog):
    """Окно просмотра зафиксированных остатков на указанную дату."""

    def __init__(self, master):
        super().__init__(master)
        self.title("Просмотр остатков на дату")
        self.geometry("750x520")
        self.minsize(600, 400)

        # ─── Верхняя панель: дата + кнопка ─────────────────────────
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=15, pady=(15, 5))

        ctk.CTkLabel(top, text="Дата:", font=("Arial", 13)).pack(side="left")

        today_str = date.today().strftime(DATE_FMT)
        self.date_picker = DatePicker(top, width=120)
        self.date_picker.pack(side="left", padx=(8, 12))
        self.date_picker.insert(0, today_str)

        ctk.CTkButton(top, text="Показать", command=self.load_fixation,
                      fg_color="#4472C4", hover_color="#2e5aa8").pack(side="left")
        ctk.CTkButton(top, text="Выгрузить в Excel",
                      command=self.export_excel,
                      fg_color="#4472C4", hover_color="#2e5aa8").pack(side="right")

        # ─── Подпись результата ────────────────────────────────────
        self.info_label = ctk.CTkLabel(self, text="", font=("Arial", 12))
        self.info_label.pack(fill="x", padx=15, pady=(0, 5))

        # ─── Таблица ──────────────────────────────────────────────
        tree_frame = ctk.CTkFrame(self)
        tree_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        cols = ("type", "name", "quantity", "description")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        self.tree.heading("type",        text="Тип")
        self.tree.heading("name",        text="Название")
        self.tree.heading("quantity",    text="Количество")
        self.tree.heading("description", text="Описание")

        self.tree.column("type",        width=100, minwidth=80,  stretch=False)
        self.tree.column("name",        width=200, minwidth=120, stretch=True)
        self.tree.column("quantity",    width=110, minwidth=80,  stretch=False, anchor="center")
        self.tree.column("description", width=250, minwidth=100, stretch=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

    # ─────────────────────────────────────────────────────────────

    def load_fixation(self):
        """Загрузить фиксацию на выбранную дату."""
        raw = self.date_picker.get().strip()
        try:
            target = datetime.strptime(raw, DATE_FMT).date()
        except ValueError:
            messagebox.showwarning("Ошибка", "Формат даты: ДД.ММ.ГГГГ", parent=self)
            return

        fixations = get_all_fixations()

        chosen = None
        for f in fixations:
            f_date = datetime.strptime(f["date"], DATE_DB).date()
            if f_date == target:
                chosen = f
                break
            if f_date < target:
                chosen = f
                break

        self.tree.delete(*self.tree.get_children())

        if chosen is None:
            self.info_label.configure(text="Нет зафиксированных остатков до этой даты.")
            return

        fix_date = datetime.strptime(chosen["date"], DATE_DB).strftime(DATE_FMT)
        fix_time = chosen.get("timestamp", "")[:19].replace("T", " ")
        self.info_label.configure(
            text=f"Фиксация от {fix_date}  ({fix_time})"
        )

        for d in chosen.get("discs", []):
            self.tree.insert("", "end", values=(
                "Носитель", d["name"], d["quantity"], d.get("description", "")
            ))
        for b in chosen.get("boxes", []):
            self.tree.insert("", "end", values=(
                "Коробка", b["name"], b["quantity"], b.get("description", "")
            ))

    def export_excel(self):
        """Выгрузить таблицу остатков в Excel."""
        rows = []
        for item_id in self.tree.get_children():
            item = self.tree.item(item_id)
            vals = item['values']
            rows.append([vals[0], vals[1], vals[2], vals[3]])

        if not rows:
            messagebox.showinfo("Нет данных", "Нет данных для выгрузки.", parent=self)
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Остатки"

        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        align_c = Alignment(horizontal='center', vertical='center')
        align_l = Alignment(horizontal='left', vertical='center')

        headers = ["Тип", "Название", "Количество", "Описание"]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_c
            cell.border = border

        for ri, vals in enumerate(rows, 2):
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = border
                cell.alignment = align_c if ci == 3 else align_l

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 40

        suggested = f"Остатки от {self.date_picker.get().replace('.','-')}.xlsx"
        file_path = ctk.filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=suggested,
            parent=self,
        )
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Успех", f"Файл сохранён:\n{file_path}", parent=self)
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}", parent=self)

    def on_close(self):
        super().on_close()
