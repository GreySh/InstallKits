"""
Окно отчетов.
"""

from tkinter import messagebox
import customtkinter as ctk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import os
from data import get_all_products, get_product_components, get_stock_disc_quantity, get_stock_box_quantity, get_all_stock_discs, get_all_stock_boxes
from ui.dialogs.base_dialog import BaseDialog


class ReportWindow(BaseDialog):
    def get_default_geometry(self):
        return "600x500"
    
    def __init__(self, master):
        super().__init__(master)
        
        self.title("Отчеты")
        
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        # Заголовок
        ctk.CTkLabel(self, text="Отчет по остаткам", font=("Arial", 16, "bold")).grid(row=0, column=0, padx=10, pady=10, sticky="w")
        
        # Таблица
        self.report_tree = ttk.Treeview(self, columns=("available",), show="tree headings")
        self.report_tree.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        
        self.report_tree.heading("#0", text="Продукт/Носитель/Коробка")
        self.report_tree.heading("available", text="Доступно/Остаток")
        
        # Настройка ширин колонок
        self.report_tree.column("#0", width=200, anchor="w")
        self.report_tree.column("available", width=50, anchor="w")
        
        # Кнопка экспорта
        self.export_button = ctk.CTkButton(self, text="Экспортировать в Excel", command=self.export_excel)
        self.export_button.grid(row=2, column=0, padx=10, pady=10, sticky="e")
        
        # Кнопка закрытия
        self.close_button = ctk.CTkButton(self, text="Закрыть", command=self.destroy)
        self.close_button.grid(row=2, column=0, padx=10, pady=10, sticky="w")
        
        # Загрузить отчет
        self.load_report()
    
    def load_report(self):
        """Загрузить отчет."""
        # Очистить таблицу
        for item in self.report_tree.get_children():
            self.report_tree.delete(item)
        
        # Загрузить продукты
        products = get_all_products()
        for product in products:
            product_id = product.doc_id
            name = product['name']
            components = get_product_components(product_id)
            
            # Определить минимальное доступное количество комплектов
            min_quantity = float('inf')
            for comp in components:
                disc_qty = get_stock_disc_quantity(comp['disc_id']) if comp['disc_id'] else float('inf')
                box_qty = get_stock_box_quantity(comp['box_id']) if comp['box_id'] else float('inf')
                
                if comp['disc_id'] and comp['disc_quantity'] > 0:
                    disc_qty //= comp['disc_quantity']
                if comp['box_id'] and comp['box_quantity'] > 0:
                    box_qty //= comp['box_quantity']
                
                min_quantity = min(min_quantity, disc_qty, box_qty)
            
            if min_quantity == float('inf'):
                min_quantity = 0
            
            self.report_tree.insert("", "end", text=name, values=(int(min_quantity),))
        
        # Загрузить носители
        discs = get_all_stock_discs()
        for disc in discs:
            self.report_tree.insert("", "end", text=f"[Носитель] {disc['disc_name']}", values=(disc['quantity'],))
        
        # Загрузить коробки
        boxes = get_all_stock_boxes()
        for box in boxes:
            self.report_tree.insert("", "end", text=f"[Коробка] {box['box_name']}", values=(box['quantity'],))
    
    def export_excel(self):
        """Экспортировать в Excel."""
        # Создать новую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Остатки комплектов"
        
        # Заголовки
        ws['A1'] = "Продукт/Носитель/Коробка"
        ws['B1'] = "Доступно/Остаток"
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 11

        # Настройка выравнивания для колонки B (по левому краю)
        left_alignment = Alignment(horizontal='left')
        
        # Дата отчета в ячейке A2
        ws['C1'] = "Дата отчета:"
        ws['D1'] = datetime.now().strftime('%d.%m.%Y')
        
        # Данные - продукты
        products = get_all_products()
        row = 2
        for product in products:
            product_id = product.doc_id
            name = product['name']
            components = get_product_components(product_id)
            
            # Определить минимальное доступное количество комплектов
            min_quantity = float('inf')
            for comp in components:
                disc_qty = get_stock_disc_quantity(comp['disc_id']) if comp['disc_id'] else float('inf')
                box_qty = get_stock_box_quantity(comp['box_id']) if comp['box_id'] else float('inf')
                
                if comp['disc_id'] and comp['disc_quantity'] > 0:
                    disc_qty //= comp['disc_quantity']
                if comp['box_id'] and comp['box_quantity'] > 0:
                    box_qty //= comp['box_quantity']
                
                min_quantity = min(min_quantity, disc_qty, box_qty)
            
            if min_quantity == float('inf'):
                min_quantity = 0
            
            ws[f'A{row}'] = name
            ws[f'B{row}'] = int(min_quantity)
            ws[f'B{row}'].alignment = left_alignment
            row += 1
        
        # Данные - носители
        discs = get_all_stock_discs()
        for disc in discs:
            ws[f'A{row}'] = f"[Носитель] {disc['disc_name']}"
            ws[f'B{row}'] = disc['quantity']
            ws[f'B{row}'].alignment = left_alignment
            row += 1
        
        # Данные - коробки
        boxes = get_all_stock_boxes()
        for box in boxes:
            ws[f'A{row}'] = f"[Коробка] {box['box_name']}"
            ws[f'B{row}'] = box['quantity']
            ws[f'B{row}'].alignment = left_alignment
            row += 1
        
        # Предложить имя файла
        default_filename = f"Отчет по остаткам ИК_{datetime.now().strftime('%Y%m%d')}.xlsx"

        # Открыть диалог сохранения
        file_path = ctk.filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_filename
        )
        
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Успех", f"Отчет сохранен в {file_path}")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {str(e)}")
