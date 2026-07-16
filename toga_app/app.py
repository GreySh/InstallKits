"""
Главное приложение Toga для учёта инсталляционных комплектов.
"""

from datetime import datetime
import asyncio

import toga
from toga.style import Pack
from toga.style.pack import COLUMN, ROW
from toga.sources.columns import AccessorColumn

# ── Патч бэкенда: поддержка индивидуальных ширин и выравнивания колонок ──
# В установленной версии Toga Table-бэкенд (winforms) игнорирует ширины
# колонок и делает их равными, а также не поддерживает выравнивание.
# Патчим _resize_columns, чтобы учитывать атрибуты .width и .align,
# заданные на объекте колонки (без .width = flex-колонка).
try:
    import toga_winforms.widgets.table as _tbl
    _WinFormsTableView = _tbl.Table
    _WinForms = _tbl.WinForms

    def _resize_columns_with_widths(self):
        num_cols = len(self.native.Columns)
        if num_cols == 0:
            return
        toga_cols = self.interface._columns
        total = self.native.ClientSize.Width
        fixed = 0
        flex_idx = []
        for i, tc in enumerate(toga_cols):
            w = getattr(tc, "width", None)
            if w:
                fixed += int(w)
            else:
                flex_idx.append(i)
        remaining = max(0, total - fixed - 4)
        each = remaining // len(flex_idx) if flex_idx else 0
        align_map = {
            "left": _WinForms.HorizontalAlignment.Left,
            "center": _WinForms.HorizontalAlignment.Center,
            "right": _WinForms.HorizontalAlignment.Right,
        }
        for i, tc in enumerate(toga_cols):
            w = getattr(tc, "width", None)
            size = int(w) if w else each
            self.native.Columns[i].Width = size
            al = getattr(tc, "align", None)
            if al in align_map:
                self.native.Columns[i].TextAlign = align_map[al]

    _WinFormsTableView._resize_columns = _resize_columns_with_widths
except Exception:
    pass

from data import (
    get_all_products, get_product_components, delete_product,
    get_product_available_quantity, dispatch_product, dispatch_products_batch,
    get_all_stock_discs, get_all_stock_boxes, delete_disc, delete_box,
    get_operations,
)
from settings_manager import get_db_path, set_db_path
try:
    from .dialogs import (
        AddComponentWindow, AdjustStockWindow, WriteOffWindow,
        AddEditProductWindow, apply_geometry, save_geometry,
    )
except ImportError:
    from dialogs import (
        AddComponentWindow, AdjustStockWindow, WriteOffWindow,
        AddEditProductWindow, apply_geometry, save_geometry,
    )

NAV = [
    ("Списание", "dispatch"),
    ("Состав ИК", "products"),
    ("Остатки", "stock"),
    ("Операции", "operations"),
    ("Настройки", "settings"),
]

BLUE = "#366092"
GREY = "#555555"


def ns(**kw):
    return dict(kw)


class InstallKitsApp(toga.App):
    def startup(self):
        self.main_window = toga.MainWindow(
            title="Учёт инсталляционных комплектов", resizable=True)
        self.content_box = toga.Box(style=Pack(direction=COLUMN, margin=10, flex=1))
        self.nav_box = toga.Box(style=Pack(direction=COLUMN, width=170, margin=6))

        self.nav_buttons = {}
        for label, key in NAV:
            btn = toga.Button(label, on_press=lambda w, k=key: self.show(k),
                              style=Pack(margin=4, background_color=BLUE, color="#ffffff", font_size=12))
            self.nav_buttons[key] = btn
            self.nav_box.add(btn)

        main = toga.Box(style=Pack(direction=ROW))
        main.add(self.nav_box)
        main.add(self.content_box)

        self.main_window.content = main
        apply_geometry(self.main_window, "MainWindow", default_size=(1024, 700))
        self.main_window.show()
        self.main_window.on_resize = self._on_main_geometry_change
        self._main_shown = True
        self.show("dispatch")

    # ─── Навигация ──────────────────────────────────────────────

    def show(self, key):
        for k, b in self.nav_buttons.items():
            b.style.background_color = BLUE if k == key else "#24456b"
        self.content_box.clear()
        builder = getattr(self, f"_build_{key}")
        widget = builder()
        self.content_box.add(widget)
        # WinForms-бэкенд не привязывает данные таблицы при создании —
        # повторно присваиваем data, чтобы строки отобразились.
        self._rebind_tables(widget)
        if getattr(self, "_main_shown", False):
            self._save_main_geometry()

    def _save_main_geometry(self):
        try:
            save_geometry("MainWindow", self.main_window.size,
                         self.main_window.position)
        except Exception:
            pass

    def _on_main_geometry_change(self, window=None):
        if getattr(self, "_main_shown", False):
            self._save_main_geometry()

    def _rebind_tables(self, widget):
        from toga.widgets.table import Table
        if isinstance(widget, Table):
            try:
                widget.data = widget.data
            except Exception:
                pass
        for child in (getattr(widget, "children", None) or []):
            self._rebind_tables(child)

    # ─── Списание ───────────────────────────────────────────────

    def _heading(self, text):
        return toga.Label(text, style=Pack(margin=4, font_size=18, font_weight="bold"))

    def _build_dispatch(self):
        box = toga.Box(style=Pack(direction=COLUMN, flex=1, margin=4))

        heading = toga.Box(style=Pack(direction=ROW, margin=4))
        heading.add(toga.Box(style=Pack(flex=1)))
        heading.add(toga.Label("Списание инсталляционных комплектов",
                            style=Pack(margin=4, font_size=18, font_weight="bold")))
        heading.add(toga.Box(style=Pack(flex=1)))
        box.add(heading)

        header_row = toga.Box(style=Pack(direction=ROW, margin=4, align_items="center"))
        header_row.add(toga.Label("Название", style=Pack(flex=2, margin=4, font_size=12, font_weight="bold")))
        header_row.add(toga.Label("Описание", style=Pack(flex=3, margin=4, color=GREY, font_size=12, font_weight="bold")))
        header_row.add(toga.Label("Доступно", style=Pack(width=110, margin=4, font_size=12, font_weight="bold", text_align="center")))
        header_row.add(toga.Label("Списать", style=Pack(width=80, margin=4, font_size=12, font_weight="bold", text_align="right")))
        box.add(header_row)

        rows_box = toga.Box(style=Pack(direction=COLUMN, flex=1, margin=4))
        self.dispatch_rows = []
        products = get_all_products()
        if not products:
            rows_box.add(toga.Label("Нет комплектов", style=Pack(margin=8, color=GREY, font_size=12)))
        for p in products:
            pid = p.doc_id
            available = get_product_available_quantity(pid)
            row = toga.Box(style=Pack(direction=ROW, margin=4, align_items="center"))
            row.add(toga.Label(p["name"], style=Pack(flex=2, margin=4, font_size=12, font_weight="bold")))
            row.add(toga.Label(p.get("description", "") or "—", style=Pack(flex=3, margin=4, color=GREY, font_size=12)))
            row.add(toga.Label(f"{available}", style=Pack(width=110, margin=4, color="#27ae60", font_size=12, text_align="center")))
            qty = toga.NumberInput(min=0, value=0, style=Pack(width=50, margin=4, font_size=12, text_align="right"))
            row.add(qty)
            rows_box.add(row)
            self.dispatch_rows.append((p, available, qty))

        sc = toga.ScrollContainer(vertical=True, horizontal=False, style=Pack(flex=1, margin=4))
        sc.content = rows_box
        box.add(sc)

        bottom = toga.Box(style=Pack(direction=ROW, margin=4, align_items="center"))
        bottom.add(toga.Box(style=Pack(flex=1)))
        bottom.add(toga.Label("Дата отгрузки:", style=Pack(margin=4, font_size=12)))
        self.dispatch_date = toga.DateInput(value=datetime.now().date(),
                                            style=Pack(width=160, margin=4, font_size=12))
        bottom.add(self.dispatch_date)
        bottom.add(toga.Button("Списать выбранное", on_press=self._dispatch_selected,
                                style=Pack(margin=8, background_color="#c0392b", color="#ffffff", font_size=12)))
        box.add(bottom)
        return box

    def _comps_text(self, pid):
        parts = []
        for c in get_product_components(pid):
            if c.get("disc_id"):
                d = get_disc_by_id_safe(c["disc_id"])
                if d:
                    parts.append(f"{d['name']}×{c.get('disc_quantity', 0)}")
            if c.get("box_id"):
                b = get_box_by_id_safe(c["box_id"])
                if b:
                    parts.append(f"{b['name']}×{c.get('box_quantity', 0)}")
        return "; ".join(parts) if parts else "—"

    async def _dispatch_selected(self, widget):
        dispatch_date = self._date_value()
        if dispatch_date is None:
            await self.main_window.dialog(toga.ErrorDialog("Ошибка", "Укажите дату отгрузки"))
            return
        items = []
        for p, available, qty_widget in self.dispatch_rows:
            q = int(qty_widget.value or 0)
            if q > 0:
                items.append((p.doc_id, q))
        if not items:
            await self.main_window.dialog(toga.InfoDialog("Информация", "Укажите количество хотя бы для одного комплекта"))
            return
        ok, messages = dispatch_products_batch(items, dispatch_date)
        if ok:
            text = "\n".join(messages) if isinstance(messages, list) else str(messages)
            await self.main_window.dialog(toga.InfoDialog("Успех", text))
            self.show("dispatch")
        else:
            text = "\n".join(messages) if isinstance(messages, list) else str(messages)
            await self.main_window.dialog(toga.ErrorDialog("Ошибка", text))

    def _date_value(self):
        d = self.dispatch_date.value
        return d.strftime("%Y-%m-%d") if d else None

    # ─── Состав ИК (продукты) ───────────────────────────────────

    def _build_products(self):
        box = toga.Box(style=Pack(direction=COLUMN, flex=1, margin=4))
        box.add(self._heading("Комплекты"))
        controls = toga.Box(style=Pack(direction=ROW, margin=4))
        controls.add(toga.Button("Добавить", on_press=lambda w: AddEditProductWindow(
            on_save=lambda: self.show("products")).show(), style=Pack(margin=4, background_color="#27ae60", color="#ffffff", font_size=12)))
        controls.add(toga.Button("Редактировать", on_press=self._edit_product,
                                style=Pack(margin=4, background_color=BLUE, color="#ffffff", font_size=12)))
        controls.add(toga.Button("Удалить", on_press=self._delete_product,
                                style=Pack(margin=4, background_color="#c0392b", color="#ffffff", font_size=12)))
        box.add(controls)

        data = [ns(id=p.doc_id, name=p["name"], available=get_product_available_quantity(p.doc_id),
                   components=self._comps_text(p.doc_id)) for p in get_all_products()]
        table = toga.Table(
            columns=[
                self._op_col("Название", "name", 200),
                self._op_col("Доступно", "available", 100, "center"),
                self._op_col("Состав", "components", None),
            ],
            data=data, style=Pack(flex=1, margin=4, font_size=12),
            on_select=self._on_select_products,
        )
        self.products_table = table
        box.add(table)
        return box

    def _on_select_products(self, widget):
        self.products_selected = widget.selection

    async def _edit_product(self, widget):
        row = getattr(self, "products_selected", None)
        if row is None:
            await self.main_window.dialog(toga.InfoDialog("Информация", "Выберите комплект для редактирования"))
            return
        AddEditProductWindow(product_id=row.id, on_save=lambda: self.show("products")).show()

    async def _delete_product(self, widget):
        row = getattr(self, "products_selected", None)
        if row is None:
            await self.main_window.dialog(toga.InfoDialog("Информация", "Выберите комплект для удаления"))
            return
        if await self.main_window.dialog(toga.QuestionDialog("Подтверждение", f'Удалить комплект "{row.name}"?')):
            delete_product(row.id)
            self.show("products")

    # ─── Остатки ────────────────────────────────────────────────

    def _build_stock(self):
        self.stock_mode = "disc"
        box = toga.Box(style=Pack(direction=COLUMN, flex=1, margin=4))
        header = toga.Box(style=Pack(direction=ROW, margin=4, align_items="center"))
        header.add(self._heading("Склад"))
        header.add(toga.Box(style=Pack(flex=1)))
        header.add(toga.Button("Выгрузить остатки в Excel", on_press=self._export_report,
                                style=Pack(margin=4, background_color="#4472C4", color="#ffffff", font_size=12)))
        box.add(header)
        subtabs = toga.Box(style=Pack(direction=ROW, margin=4))
        self.subtab_disc = toga.Button("Носители", on_press=lambda w: self._stock_toggle("disc"),
                                       style=Pack(margin=4, background_color=BLUE, color="#ffffff", font_size=12))
        self.subtab_box = toga.Button("Коробки", on_press=lambda w: self._stock_toggle("box"),
                                      style=Pack(margin=4, background_color="#24456b", color="#ffffff", font_size=12))
        subtabs.add(self.subtab_disc)
        subtabs.add(self.subtab_box)
        box.add(subtabs)
        top = toga.Box(style=Pack(direction=ROW, margin=4))
        top.add(toga.Button("Добавить", on_press=self._stock_add,
                            style=Pack(margin=4, background_color="#27ae60", color="#ffffff", font_size=12)))
        top.add(toga.Button("Корректировка", on_press=self._stock_adjust,
                            style=Pack(margin=4, background_color=BLUE, color="#ffffff", font_size=12)))
        top.add(toga.Button("Брак", on_press=self._stock_writeoff,
                            style=Pack(margin=4, background_color="#c0392b", color="#ffffff", font_size=12)))
        top.add(toga.Button("Удалить", on_press=self._stock_delete,
                            style=Pack(margin=4, background_color="#c0392b", color="#ffffff", font_size=12)))
        top.add(toga.Button("Новый компонент", on_press=lambda w: AddComponentWindow(
            self.stock_mode, on_save=lambda: self.show("stock")).show(),
            style=Pack(margin=4, background_color="#27ae60", color="#ffffff", font_size=12)))
        box.add(top)

        self.stock_inner = toga.Box(style=Pack(direction=COLUMN, flex=1, margin=4))
        box.add(self.stock_inner)
        self._stock_rebuild()
        return box

    def _stock_toggle(self, mode=None):
        if mode is None:
            mode = self.stock_mode
        self.stock_mode = mode
        self.subtab_disc.style.background_color = BLUE if mode == "disc" else "#24456b"
        self.subtab_box.style.background_color = BLUE if mode == "box" else "#24456b"
        self._stock_rebuild()

    def _stock_rebuild(self):
        self.stock_inner.clear()
        if self.stock_mode == "disc":
            data = [ns(id=d["disc_id"], name=d["disc_name"], quantity=d["quantity"],
                      description=d.get("description", ""))
                    for d in get_all_stock_discs()]
        else:
            data = [ns(id=b["box_id"], name=b["box_name"], quantity=b["quantity"],
                      description=b.get("description", ""))
                    for b in get_all_stock_boxes()]
        table = toga.Table(
            columns=[
                self._op_col("Название", "name", 200),
                self._op_col("Кол-во", "quantity", 80, "center"),
                self._op_col("Описание", "description", None),
            ],
            data=data, style=Pack(flex=1, margin=4, font_size=12),
            on_select=self._on_select_stock,
        )
        self.stock_inner.add(table)
        table.data = data

    def _on_select_stock(self, widget):
        sel = widget.selection
        if sel is not None:
            self.stock_selected = {"id": sel.id, "name": sel.name, "quantity": sel.quantity}
        else:
            self.stock_selected = None

    async def _stock_add(self, widget):
        item = getattr(self, "stock_selected", None)
        if item is None:
            await self.main_window.dialog(toga.InfoDialog("Информация", "Выберите запись"))
            return
        AdjustStockWindow(self.stock_mode, item, on_save=lambda: self.show("stock"),
                          fixed_op="add").show()

    async def _stock_adjust(self, widget):
        item = getattr(self, "stock_selected", None)
        if item is None:
            await self.main_window.dialog(toga.InfoDialog("Информация", "Выберите запись"))
            return
        AdjustStockWindow(self.stock_mode, item, on_save=lambda: self.show("stock"),
                          allow_rename=True).show()

    async def _stock_writeoff(self, widget):
        item = getattr(self, "stock_selected", None)
        if item is None:
            await self.main_window.dialog(toga.InfoDialog("Информация", "Выберите запись"))
            return
        WriteOffWindow(self.stock_mode, item, on_save=lambda: self.show("stock")).show()

    async def _stock_delete(self, widget):
        item = getattr(self, "stock_selected", None)
        if item is None:
            await self.main_window.dialog(toga.InfoDialog("Информация", "Выберите запись"))
            return
        if await self.main_window.dialog(toga.QuestionDialog("Подтверждение", f'Удалить "{item["name"]}"?')):
            if self.stock_mode == "disc":
                delete_disc(item["id"])
            else:
                delete_box(item["id"])
            self.show("stock")

    # ─── Отчёты ─────────────────────────────────────────────────

    def _build_web_report(self, title, subtitle, sections, widths, wrap=False):
        """Построить книгу в формате веб-версии.

        sections — список (заголовок_секции, заголовки, строки, left_columns),
        где left_columns — множество 0-based индексов столбцов с выравниванием влево.
        Возвращает (wb, ws).
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        HEADER_FILL = PatternFill("solid", fgColor="366092")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
        TITLE_FONT = Font(bold=True, size=14)
        SUBTITLE_FONT = Font(italic=True, size=10)
        SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")
        SECTION_FILL = PatternFill("solid", fgColor="4472C4")
        BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        CENTER = Alignment(horizontal='center', vertical='center')
        LEFT = Alignment(horizontal='left', vertical='center')
        CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
        LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_c = CENTER_WRAP if wrap else CENTER
        align_l = LEFT_WRAP if wrap else LEFT

        wb = Workbook()
        ws = wb.active
        ws.title = title
        ncols = max((len(h) for _, h, _, _ in sections), default=1)
        last_col = get_column_letter(ncols)
        ws['A1'] = title
        ws['A1'].font = TITLE_FONT
        ws.merge_cells(f'A1:{last_col}1')
        ws['A1'].alignment = align_c
        ws['A2'] = subtitle
        ws['A2'].font = SUBTITLE_FONT
        ws.merge_cells(f'A2:{last_col}2')
        row = 4
        first_data_row = None
        for sec_title, headers, rows, left_columns in sections:
            if left_columns is None:
                left_columns = {0}
            ws[f'A{row}'] = sec_title
            ws[f'A{row}'].font = SECTION_FONT
            ws[f'A{row}'].fill = SECTION_FILL
            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).fill = SECTION_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            row += 1
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=ci)
                cell.value = h
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = align_c
                cell.border = BORDER
            row += 1
            if first_data_row is None:
                first_data_row = row
            for ri, vals in enumerate(rows):
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=ci)
                    cell.value = v
                    cell.alignment = align_l if (ci - 1) in left_columns else align_c
                    cell.border = BORDER
                    if ri % 2 == 1:
                        cell.fill = PatternFill("solid", fgColor="F2F6FA")
                row += 1
            row += 1
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        if first_data_row is not None:
            ws.freeze_panes = ws.cell(row=first_data_row, column=1)
        return wb, ws

    async def _export_report(self, widget):
        product_rows = [[p["name"], get_product_available_quantity(p.doc_id)]
                        for p in get_all_products()]
        disc_rows = [[d["disc_name"], d["quantity"]] for d in get_all_stock_discs()]
        box_rows = [[b["box_name"], b["quantity"]] for b in get_all_stock_boxes()]
        wb, ws = self._build_web_report(
            "ОТЧЕТ ПО ОСТАТКАМ",
            f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}",
            [
                ("ПРОДУКТЫ", ["Продукт", "Доступно (комплектов)"], product_rows, {0}),
                ("НОСИТЕЛИ", ["Носитель", "Остаток"], disc_rows, {0}),
                ("КОРОБКИ", ["Коробка", "Остаток"], box_rows, {0}),
            ],
            widths={'A': 35, 'B': 25},
        )
        suggested = f"Отчет по остаткам ИК_{datetime.now().strftime('%Y%m%d')}.xlsx"
        try:
            result = await self.main_window.save_file_dialog(
                "Сохранить отчёт по остаткам",
                suggested_filename=suggested,
                file_types=["xlsx"],
            )
        except Exception:
            return
        if not result:
            return
        fname = result[0] if isinstance(result, (list, tuple)) else result
        if not fname:
            return
        try:
            wb.save(str(fname))
            await self.main_window.dialog(toga.InfoDialog("Успех", f"Файл сохранён: {fname}"))
        except Exception as e:
            await self.main_window.dialog(toga.ErrorDialog("Ошибка", f"Не удалось сохранить: {e}"))

    # ─── Операции ───────────────────────────────────────────────

    OP_MAP = {
        "add_disc": "Добавлены носители",
        "add_box": "Добавлены коробки",
        "adjust_stock": "Корректировка остатка",
        "dispatch": "Списание комплекта",
        "write_off": "Списание по браку",
    }

    def _op_col(self, heading, accessor, width, align=None):
        col = AccessorColumn(heading=heading, accessor=accessor)
        if width is not None:
            col.width = width
        if align is not None:
            col.align = align
        return col

    def _build_operations(self):
        box = toga.Box(style=Pack(direction=COLUMN, flex=1, margin=4))
        box.add(self._heading("История операций"))
        filters = toga.Box(style=Pack(direction=ROW, margin=4, align_items="center"))
        filters.add(toga.Label("С:", style=Pack(margin=4, font_size=12)))
        self.op_start = toga.DateInput(style=Pack(width=160, margin=4, font_size=12))
        filters.add(self.op_start)
        filters.add(toga.Label("По:", style=Pack(margin=4, font_size=12)))
        self.op_end = toga.DateInput(style=Pack(width=160, margin=4, font_size=12))
        filters.add(self.op_end)
        # По умолчанию «С:» — дата первой операции в базе
        try:
            first_op = min((op.get("date", "") for op in get_operations() if op.get("date")), default="")
            if first_op:
                self.op_start.value = datetime.strptime(first_op[:10], "%Y-%m-%d").date()
        except Exception:
            pass
        self.op_type = toga.Selection(items=["Все"] + list(self.OP_MAP.values()),
                                      style=Pack(width=200, margin=4, font_size=12))
        filters.add(self.op_type)
        filters.add(toga.Button("Показать", on_press=lambda w: self._op_refresh(),
                                style=Pack(margin=4, background_color=BLUE, color="#ffffff", font_size=12)))
        filters.add(toga.Button("Excel", on_press=self._export_operations,
                                style=Pack(margin=4, background_color="#4472C4", color="#ffffff", font_size=12)))
        box.add(filters)

        self.op_table = toga.Table(
            columns=[
                self._op_col("Дата", "date", 170, "left"),
                self._op_col("Тип", "type", 200),
                self._op_col("Продукт", "product", 180),
                self._op_col("Кол-во", "quantity", 80, "center"),
                self._op_col("Детали", "details", None),
            ],
            data=[], style=Pack(flex=1, margin=4, font_size=12),
        )
        box.add(self.op_table)
        self._op_refresh()
        return box

    def _op_refresh(self):
        from data import get_product_by_id
        reverse = {v: k for k, v in self.OP_MAP.items()}
        type_key = reverse.get(self.op_type.value)
        start_dt = self.op_start.value.strftime("%Y-%m-%d") if self.op_start.value else None
        end_dt = self.op_end.value.strftime("%Y-%m-%d") if self.op_end.value else None

        data = []
        for op in get_operations():
            date = op.get("date", "")
            if start_dt and date[:10] < start_dt:
                continue
            if end_dt and date[:10] > end_dt:
                continue
            if type_key and op.get("operation_type") != type_key:
                continue
            pid = op.get("product_id")
            pname = "-"
            if pid:
                pr = get_product_by_id(pid)
                pname = pr["name"] if pr else "Неизвестно"
            details = op.get("details", {})
            dtext = ""
            if isinstance(details, dict):
                comps = details.get("components", [])
                if comps:
                    parts = [f"{c.get('type','').upper()}: {c.get('name','')} ({c.get('quantity', c.get('total_quantity',''))} шт.)"
                             for c in comps]
                    dtext = "; ".join(parts)
                if details.get("reason"):
                    dtext += f" | Причина: {details['reason']}"
            data.append(ns(date=date, type=self.OP_MAP.get(op.get("operation_type"), ""),
                            product=pname, quantity=op.get("quantity", ""), details=dtext))
        self.op_table.data = data

    async def _export_operations(self, widget):
        from data import get_product_by_id
        reverse = {v: k for k, v in self.OP_MAP.items()}
        type_key = reverse.get(self.op_type.value)
        start_dt = self.op_start.value.strftime("%Y-%m-%d") if self.op_start.value else None
        end_dt = self.op_end.value.strftime("%Y-%m-%d") if self.op_end.value else None
        rows = []
        for op in get_operations():
            date = op.get("date", "")
            if start_dt and date[:10] < start_dt:
                continue
            if end_dt and date[:10] > end_dt:
                continue
            if type_key and op.get("operation_type") != type_key:
                continue
            pid = op.get("product_id")
            pname = "-"
            if pid:
                pr = get_product_by_id(pid)
                pname = pr["name"] if pr else "Неизвестно"
            details = op.get("details", {})
            dtext = ""
            if isinstance(details, dict):
                comps = details.get("components", [])
                if comps:
                    parts = [f"{c.get('type','').upper()}: {c.get('name','')} ({c.get('quantity', c.get('total_quantity',''))} шт.)" for c in comps]
                    dtext = "; ".join(parts)
                if details.get("reason"):
                    dtext += f" | Причина: {details['reason']}"
            rows.append([date, self.OP_MAP.get(op.get("operation_type"), ""), pname,
                         op.get("quantity", ""), dtext])
        wb, ws = self._build_web_report(
            "ОТЧЕТ О ОПЕРАЦИЯХ",
            f"Период: {self.op_start.value or 'Начало'} - {self.op_end.value or 'Текущее время'}",
            [("ОПЕРАЦИИ", ["Дата и время", "Тип операции", "Продукт", "Количество", "Детали"],
              rows, {0, 1, 2, 4})],
            widths={'A': 20, 'B': 22, 'C': 25, 'D': 12, 'E': 40},
            wrap=True,
        )
        suggested = f"Операции_{datetime.now().strftime('%Y%m%d')}.xlsx"
        try:
            result = await self.main_window.save_file_dialog(
                "Сохранить отчёт по операциям",
                suggested_filename=suggested,
                file_types=["xlsx"],
            )
        except Exception:
            return
        if not result:
            return
        fname = result[0] if isinstance(result, (list, tuple)) else result
        if not fname:
            return
        try:
            wb.save(str(fname))
            await self.main_window.dialog(toga.InfoDialog("Успех", f"Файл сохранён: {fname}"))
        except Exception as e:
            await self.main_window.dialog(toga.ErrorDialog("Ошибка", f"Не удалось сохранить: {e}"))

    # ─── Настройки ──────────────────────────────────────────────

    def _build_settings(self):
        box = toga.Box(style=Pack(direction=COLUMN, flex=1, margin=10))
        box.add(self._heading("Настройки"))
        box.add(toga.Label("Настройки базы данных", style=Pack(margin=4, font_size=12, font_weight="bold")))
        self.db_path = toga.TextInput(value=get_db_path(), style=Pack(flex=1, margin=4, font_size=12))
        row = toga.Box(style=Pack(direction=ROW, margin=4, align_items="center"))
        row.add(toga.Label("База данных:", style=Pack(width=140, margin=4, font_size=12)))
        row.add(self.db_path)
        row.add(toga.Button("Выбрать…", on_press=self._select_db_path,
                            style=Pack(width=110, margin=4, background_color=BLUE, color="#ffffff", font_size=11)))
        box.add(row)
        actions = toga.Box(style=Pack(direction=ROW, margin=4))
        actions.add(toga.Box(style=Pack(flex=1)))
        actions.add(toga.Button("Сохранить", on_press=self._save_settings,
                                style=Pack(width=120, margin=4, background_color="#27ae60", color="#ffffff", font_size=11)))
        box.add(actions)
        box.add(toga.Label("", style=Pack(flex=1)))
        return box

    async def _select_db_path(self, widget):
        try:
            result = await self.main_window.open_file_dialog(
                "Выберите файл базы данных",
                file_types=["json", "db"],
            )
            if not result:
                return
            path = result[0] if isinstance(result, (list, tuple)) else result
            if path:
                self.db_path.value = str(path)
        except Exception:
            pass

    async def _save_settings(self, widget):
        import os
        from data.db import close_database, get_database
        p = self.db_path.value.strip()
        if not p:
            await self.main_window.dialog(toga.ErrorDialog("Ошибка", "Укажите путь к базе данных"))
            return
        try:
            d = os.path.dirname(p)
            if d and not os.path.exists(d):
                os.makedirs(d, exist_ok=True)
            set_db_path(p)
            close_database()
            get_database()
            await self.main_window.dialog(toga.InfoDialog("Успех", "Настройки сохранены. Переключитесь на другую вкладку для обновления."))
        except Exception as e:
            await self.main_window.dialog(toga.ErrorDialog("Ошибка", f"Не удалось сохранить: {e}"))


def get_disc_by_id_safe(did):
    from data import get_disc_by_id
    return get_disc_by_id(did)


def get_box_by_id_safe(bid):
    from data import get_box_by_id
    return get_box_by_id(bid)


def main():
    return InstallKitsApp("InstallKits", "org.installkits.toga")
